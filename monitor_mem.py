import openpyxl
import datetime
import psutil
import sys
import json
import subprocess
import logging.config
from traceback import print_exc, format_exc
from zipfile import ZipFile, ZIP_DEFLATED
from os import listdir, path, _exit, makedirs, remove
from concurrent.futures import ThreadPoolExecutor
from logging import getLogger

def load_config():
    global logger
    global config
    try:
        if len(sys.argv) != 2:
            print('Please set config.json path to the argument.', file=sys.stderr)
            _exit(1)
        else:
            config_path = sys.argv[1]
            if path.exists(config_path) is False:
                print('config.json path is invalid.', file=sys.stderr)
                _exit(1)
            else:
                with open(config_path, 'r', encoding='utf-8_sig') as f:
                    config = json.load(f)
    except:
        print(print_exc(), file=sys.stderr)
        _exit(1)

    logging.config.dictConfig(config['log_config'])
    logger = getLogger('main')

def monitor_metrics():
    data = []
    logger.info('Start to get timestamp.')
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    logger.info('Complated to get timestamp.')
    logger.info('Start to create the get_metrics threads.')
    with ThreadPoolExecutor(max_workers=config['max_thread']) as executor:
        for proc in psutil.process_iter():
            executor.submit(get_metricts, data, proc, timestamp)
    logger.info('Complete to create the get_metrics threads.')
    return data

def get_metricts(data, proc, timestamp):
    pinfo = { "timestamp":  timestamp }
    pinfo.update(proc.as_dict(attrs=config['metrics']))
    pinfo['memory'] = proc.memory_info().vms
    if 'create_time' in pinfo:
        pinfo['create_time'] = datetime.datetime.utcfromtimestamp(pinfo['create_time'])
    cpu_percent = subprocess.run(
        ['powershell', '-Command', 'Get-Process -Id ' + str(pinfo['pid']) + ' | Select-Object -Expand CPU'],
        capture_output=True
    ).stdout.decode().replace('\r\n', '')
    if cpu_percent == '':
        pinfo['cpu_percent'] = 0.0
    else:
        pinfo['cpu_percent'] = float(cpu_percent)
    data.append(pinfo)

def get_disk_metrics():
    disk_info = psutil.disk_usage('/')
    return {
        "timestamp": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "disk_total": disk_info.total,
        "disk_usage": disk_info.used,
        "disk_free": disk_info.free
    }

def save_metrics(data, disk_data):
    # Initialization
    log_file_name = datetime.datetime.now().strftime('%Y%m%d') + '.xlsx'
    file_path = config['export_dir'] + '\\' + log_file_name
    if path.exists(config['export_dir']) is False:
        makedirs(config['export_dir'])
    if path.exists(file_path) is False:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet('logs')
        sheet_disk = workbook.create_sheet('disk')
        del workbook['Sheet']
        index = 1
        for key in data[0]:
            sheet.cell(row=1, column=index).value = key
            index += 1
        index = 1
        for key in disk_data:
            sheet_disk.cell(row=1, column=index).value = key
            index += 1
    else:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['logs']
        sheet_disk = workbook['disk']

    # Save metrics
    max_row = sheet.max_row
    for metrics in data:
        max_row += 1
        for i in range(sheet.max_column):
            sheet.cell(row=max_row, column=i + 1).value = metrics[sheet.cell(row=1, column=(i + 1)).value]

    # Save disk metrics
    max_row = sheet_disk.max_row + 1
    for i in range(sheet_disk.max_column):
        sheet_disk.cell(row=max_row, column=i + 1).value = disk_data[sheet_disk.cell(row=1, column=(i + 1)).value]

    workbook.save(file_path)
    workbook.close()
    return log_file_name

def archive_log(log_file_name):
    logfiles = listdir(config['export_dir'])
    if path.exists(config['archive_dir']) is False:
        makedirs(config['archive_dir'])

    for logfile in logfiles:
        if logfile != log_file_name:

            zip_file_name = logfile + '.zip'
            zip_file_path = config['archive_dir'] + '\\' + zip_file_name
            archive_log_path = config['export_dir'] + '\\' + logfile

            with ZipFile(zip_file_path, 'w') as zipfile:
                zipfile.write(archive_log_path, compress_type=ZIP_DEFLATED)
                zipfile.close()

            if path.exists(zip_file_path) is True:
                remove(archive_log_path)

if __name__ == '__main__':
    load_config()
    logger.info('Completes to load the config.json')
    try:
        logger.info('Start to get metrics.')
        data = monitor_metrics()
        logger.info('Completed to get metrics.')
        logger.info('Start to get disk metrics.')
        disk_data = get_disk_metrics()
        logger.info('Complete to get disk metrics.')
        logger.info('Start to save the log file.')
        log_file_name = save_metrics(data, disk_data)
        logger.info('Completed to save the log file.')
        logger.info('Start to archive the log file.')
        archive_log(log_file_name)
        logger.info('Completed to archive the log file.')
    except:
        print_exc()
        logger.error(format_exc())